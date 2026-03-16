"""High-level OPC-UA UDT read/write handler."""

import base64
import csv
import json
from contextlib import contextmanager
from pathlib import Path

try:
    from asyncua import ua
except ImportError:
    ua = None

from ._parser import StructuredTypeParser
from ._encoder import StructuredTypeEncoder
from ._utils import extract_fields, sanitize_name


def _excel_safe(value):
    """Strip timezone from datetime values (Excel doesn't support tzinfo)."""
    from datetime import datetime
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _flatten_to_rows(d):
    """Flatten a dict into CSV rows.

    Scalar fields and nested dicts are flattened with '_' prefix.
    List-of-dict fields are expanded: one row per list element,
    with scalar fields repeated on each row.
    """
    scalars = {}
    list_fields = {}

    for k, v in d.items():
        if isinstance(v, dict):
            for sk, sv in _flatten_to_rows(v)[0].items():
                scalars[f"{k}_{sk}"] = sv
        elif isinstance(v, list) and v and isinstance(v[0], dict):
            list_fields[k] = v
        elif isinstance(v, list):
            scalars[k] = json.dumps(v, default=str, ensure_ascii=False)
        else:
            scalars[k] = v

    if not list_fields:
        return [scalars]

    rows = []
    for field_name, items in list_fields.items():
        for item in items:
            row = dict(scalars)
            for sk, sv in _flatten_to_rows(item)[0].items():
                row[f"{field_name}_{sk}"] = sv
            rows.append(row)
    return rows


class UdtResult(list):
    """List subclass with .to_json() and .to_csv() export methods."""

    def to_json(self, filepath, indent=2):
        """Export the results to a JSON file."""
        data = [extract_fields(item) if not isinstance(item, dict) else item for item in self]
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, default=str, ensure_ascii=False)

    def to_csv(self, filepath, delimiter=';'):
        """Export the results to a CSV file.

        Nested dicts are flattened with '_' prefix.
        Lists of dicts are expanded into separate rows.
        """
        dicts = [extract_fields(item) if not isinstance(item, dict) else item for item in self]
        rows = []
        for d in dicts:
            rows.extend(_flatten_to_rows(d))
        if not rows:
            return
        fieldnames = list(dict.fromkeys(k for row in rows for k in row))
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)

    def to_excel(self, filepath):
        """Export the results to an Excel file (.xlsx).

        If the data comes from read_csv() (each element has 'type_name',
        'server_timestamp', 'values'), each CSV row becomes a separate
        sheet named '{type_name}_{server_timestamp}', containing only
        the decoded UDT values.

        Otherwise (data from read()), all values are exported into a
        single sheet.

        Requires openpyxl: pip install openpyxl
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        wb = Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)

        if self and isinstance(self[0], dict) and 'values' in self[0] and 'type_name' in self[0]:
            # Data from read_csv(): one sheet per CSV row
            for entry in self:
                type_name = entry['type_name']
                timestamp = entry.get('server_timestamp', '')
                # Remove decimals from timestamp
                timestamp = str(timestamp).split('.')[0]
                sheet_name = f"{type_name}_{timestamp}"
                # Excel sheet names max 31 chars, no special chars
                sheet_name = sheet_name[:31].replace('/', '-').replace('\\', '-')
                ws = wb.create_sheet(title=sheet_name)

                values = entry['values']
                dicts = [extract_fields(v) if not isinstance(v, dict) else v for v in values]
                rows = []
                for d in dicts:
                    rows.extend(_flatten_to_rows(d))
                if not rows:
                    continue
                fieldnames = list(dict.fromkeys(k for row in rows for k in row))
                ws.append(fieldnames)
                for row in rows:
                    ws.append([_excel_safe(row.get(k, '')) for k in fieldnames])
        else:
            # Data from read(): single sheet
            ws = wb.create_sheet(title="Data")
            dicts = [extract_fields(item) if not isinstance(item, dict) else item for item in self]
            rows = []
            for d in dicts:
                rows.extend(_flatten_to_rows(d))
            if rows:
                fieldnames = list(dict.fromkeys(k for row in rows for k in row))
                ws.append(fieldnames)
                for row in rows:
                    ws.append([_excel_safe(row.get(k, '')) for k in fieldnames])

        wb.save(filepath)


def _check_asyncua():
    if ua is None:
        raise ImportError(
            "asyncua is required for UdtHandler. "
            "Install it with: pip install uaudtcodec[opcua]"
        )


class UdtHandler:
    """High-level handler for reading/writing OPC-UA structured types."""

    def __init__(self, client, node_id, dict_node_id=None, types_cache=None):
        """
        Args:
            client: A connected asyncua.sync.Client instance.
            node_id: NodeId of the target variable (e.g. "ns=4;s=MyNode.Data").
            dict_node_id: NodeId of the type dictionary (e.g. "ns=2;i=2454").
                If None, the dictionary is auto-discovered by browsing
                the children of node i=93 (OPCBinarySchema_TypeSystem).
            types_cache: Path to a JSON cache file for type definitions.
                If the file exists, types are loaded from it (no server browsing).
                If the file does not exist, types are browsed from the server
                and saved to this file for next time.
        """
        _check_asyncua()
        self._client = client
        self._node = client.get_node(node_id)
        self._encoder_cache = {}

        if types_cache is not None and Path(types_cache).exists():
            self._load_types(types_cache)
        else:
            # Resolve the UDT type name from the node's type definition
            self._type_name = self._resolve_type_name(client, self._node)

            # Discover all type dictionaries and type nodes under i=93
            self._raw_type_dicts, self._type_node_map = self._browse_type_system(client)

            # Build encoder for the target type's dictionary
            if dict_node_id is not None:
                user_type_dict = client.get_node(dict_node_id).get_value()
                parser = StructuredTypeParser(user_type_dict)
                self._encoder = StructuredTypeEncoder(
                    parser.get_structured_types(),
                    parser.get_enumeration_types(),
                )
            else:
                dict_name = self._find_type_dict_name()
                self._encoder = self._get_encoder(dict_name)

            if types_cache is not None:
                self.save_types(types_cache)

        # Cache raw values
        self._refresh()

    def save_types(self, filepath):
        """Serialize type dictionaries and type_node_map to a JSON file.

        The saved file can be passed as types_file= to skip server browsing.

        Args:
            filepath: Output JSON file path.
        """
        data = {
            "type_name": self._type_name,
            "raw_type_dicts": {
                name: base64.b64encode(xml_bytes).decode('ascii')
                for name, xml_bytes in self._raw_type_dicts.items()
            },
            "type_node_map": {
                f"{ns_uri}|{id_str}": {"type_name": tn, "dict_name": dn}
                for (ns_uri, id_str), (tn, dn) in self._type_node_map.items()
            },
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def _load_types(self, filepath):
        """Load type dictionaries and type_node_map from a JSON file."""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        self._type_name = data["type_name"]
        self._raw_type_dicts = {
            name: base64.b64decode(b64_str)
            for name, b64_str in data["raw_type_dicts"].items()
        }
        self._type_node_map = {
            tuple(key.split("|", 1)): (v["type_name"], v["dict_name"])
            for key, v in data["type_node_map"].items()
        }
        dict_name = self._find_type_dict_name()
        self._encoder = self._get_encoder(dict_name)

    @staticmethod
    def _resolve_type_name(client, node):
        """Resolve the UDT type name from a node's data type or type definition."""
        # Try direct data type first (works for custom structured types)
        dt_node_id = node.read_data_type()
        dt_name = client.get_node(dt_node_id).read_browse_name().Name
        if dt_name not in ("Structure", "BaseDataType", "String", "ByteString"):
            return sanitize_name(dt_name)
        # Fallback: resolve via type definition → data type
        node_type_def = client.get_node(node.read_type_definition()).read_data_type()
        return sanitize_name(
            client.get_node(node_type_def).read_browse_name().Name
        )

    @staticmethod
    def _node_id_key(node_id, ns_array):
        """Build a (namespace_uri, identifier_str) key from an asyncua NodeId."""
        ns_uri = ns_array[node_id.NamespaceIndex]
        if isinstance(node_id.Identifier, int):
            id_str = f"i={node_id.Identifier}"
        else:
            id_str = f"s={node_id.Identifier}"
        return ns_uri, id_str

    @staticmethod
    def _browse_type_system(client):
        """Browse i=93 children: collect type dictionaries and map type node IDs.

        For each dict node child (type description), also follows the inverse
        HasDescription reference to find the "Default Binary" encoding node
        and maps its NodeId too (this is the encoding_node_id found in CSV exports).

        Returns:
            Tuple of:
            - raw_type_dicts: {dict_name: xml_bytes}
            - type_node_map: {(namespace_uri, identifier_str): (type_name, dict_name)}
        """
        ns_array = client.get_namespace_array()
        raw_type_dicts = {}
        type_node_map = {}
        type_system = client.get_node("i=93")

        for dict_node in type_system.get_children():
            dict_name = dict_node.read_browse_name().Name
            try:
                xml_bytes = dict_node.get_value()
                if xml_bytes is not None:
                    raw_type_dicts[dict_name] = xml_bytes
            except Exception:
                continue

            # Browse children of the dict node (type description nodes)
            for type_node in dict_node.get_children():
                try:
                    type_name = type_node.read_browse_name().Name
                    # Map the type description node itself
                    key = UdtHandler._node_id_key(type_node.nodeid, ns_array)
                    type_node_map[key] = (type_name, dict_name)

                    # Follow inverse HasDescription reference → "Default Binary" encoding node
                    for ref in type_node.get_references():
                        if ref.BrowseName.Name == "Default Binary":
                            enc_key = UdtHandler._node_id_key(ref.NodeId, ns_array)
                            type_node_map[enc_key] = (type_name, dict_name)
                except Exception:
                    continue

        return raw_type_dicts, type_node_map

    def _find_type_dict_name(self, type_name=None):
        """Find the dict_name containing the given type among the browsed dicts."""
        if type_name is None:
            type_name = self._type_name
        for name, xml_bytes in self._raw_type_dicts.items():
            try:
                parser = StructuredTypeParser(xml_bytes)
                if parser.find_structured_type(type_name):
                    return name
            except Exception:
                continue
        raise ValueError(
            f"No type dictionary found containing '{type_name}' "
            f"under node i=93. Provide dict_node_id explicitly."
        )

    def _get_encoder(self, dict_name):
        """Get (or create and cache) an encoder for the given dictionary name."""
        if dict_name not in self._encoder_cache:
            xml_bytes = self._raw_type_dicts[dict_name]
            parser = StructuredTypeParser(xml_bytes)
            self._encoder_cache[dict_name] = StructuredTypeEncoder(
                parser.get_structured_types(),
                parser.get_enumeration_types(),
            )
        return self._encoder_cache[dict_name]

    @property
    def type_node_map(self):
        """Mapping of (namespace_uri, identifier) to (type_name, dict_name).

        Contains both type description NodeIds (children of dict nodes)
        and their "Default Binary" encoding NodeIds.
        """
        return dict(self._type_node_map)

    @property
    def raw_type_dicts(self):
        """All discovered type dictionaries as {dict_name: xml_bytes}."""
        return dict(self._raw_type_dicts)

    @property
    def type_dicts(self):
        """All discovered type dictionaries, parsed.

        Returns:
            Dict of {dict_name: {"structured_types": [...], "enumeration_types": [...]}}.
        """
        result = {}
        for name, xml_bytes in self._raw_type_dicts.items():
            try:
                parser = StructuredTypeParser(xml_bytes)
                result[name] = {
                    "structured_types": parser.get_structured_types(),
                    "enumeration_types": parser.get_enumeration_types(),
                }
            except Exception:
                continue
        return result

    def _refresh(self):
        """Re-read the raw values from the server."""
        val = self._node.get_value()
        self._raw_values = val if isinstance(val, list) else [val]

    # ── Read ────────────────────────────────────────────────────────────

    def read(self, as_dict=False):
        """Read and decode all elements.

        Args:
            as_dict: If True, return list of dicts instead of typed objects.

        Returns:
            List of typed objects (or dicts). Index into the result to get
            a single element: udt.read()[2]
        """
        self._refresh()
        items = self._encoder.decode_list(self._type_name, self._raw_values)
        if as_dict:
            return UdtResult(extract_fields(item) for item in items)
        return UdtResult(items)

    def read_csv(self, filepath, as_dict=False):
        """Read and decode UDT values from a CSV export file.

        CSV columns: nodeid,datatype,opcservertimestamp,opcsourcetimestamp,statuscode,opcvalue
        opcvalue format: count;namespace_uri;encoding_node_id;base64_encoded_data

        The encoding_node_id is the "Default Binary" encoding NodeId. It is
        resolved to a type name via the type_node_map built during discovery.
        Falls back to resolving via the server using the nodeid column.

        Args:
            filepath: Path to the CSV file.
            as_dict: If True, return decoded values as dicts instead of typed objects.

        Returns:
            List of dicts, one per CSV row:
            [{"nodeid": ..., "server_timestamp": ..., "source_timestamp": ...,
              "type_name": ..., "values": [...]}, ...]
        """
        # Cache type resolution to avoid redundant lookups
        type_cache = {}

        csv.field_size_limit(2 ** 31 - 1)
        results = []
        with open(filepath, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                node_id = row['nodeid']
                parts = row['opcvalue'].split(';', 3)
                ns_uri = parts[1]
                encoding_node_id = parts[2]

                cache_key = (ns_uri, encoding_node_id)
                if cache_key not in type_cache:
                    # Try to resolve via type_node_map (encoding_node_id = Default Binary)
                    entry = self._type_node_map.get(cache_key)
                    if entry is not None:
                        type_name, dict_name = entry
                        type_name = sanitize_name(type_name)
                    else:
                        # Fallback: resolve via the server using the node_id column
                        node = self._client.get_node(node_id)
                        type_name = self._resolve_type_name(self._client, node)
                        dict_name = self._find_type_dict_name(type_name)
                    type_cache[cache_key] = (type_name, dict_name)
                type_name, dict_name = type_cache[cache_key]

                encoder = self._get_encoder(dict_name)
                unpacker = encoder._get_unpacker()


                # Each element is a separate base64 string separated by ";"
                element_parts = parts[3].rstrip(";").split(";")
                items = []
                for idx, element_b64 in enumerate(element_parts):
                    byte_buffer = base64.b64decode(element_b64)
                    item, _ = unpacker.unpack(type_name, byte_buffer, element_index=idx)
                    items.append(item)

                if as_dict:
                    items = [extract_fields(item) for item in items]

                results.append({
                    'nodeid': node_id,
                    'server_timestamp': row['opcservertimestamp'],
                    'source_timestamp': row['opcsourcetimestamp'],
                    'type_name': type_name,
                    'values': items,
                })
        return UdtResult(results)

    @property
    def count(self):
        """Number of elements in the node value."""
        return len(self._raw_values)

    @property
    def type_name(self):
        """The resolved UDT type name."""
        return self._type_name

    # ── Write ───────────────────────────────────────────────────────────

    def patch(self, index, modifications):
        """Apply partial modifications to an element and write back."""
        body = self._raw_values[index].Body if hasattr(self._raw_values[index], 'Body') else self._raw_values[index]
        new_body = self._encoder.patch(self._type_name, body, modifications)
        self._write_body(index, new_body)

    def write(self, index, instance=None, data=None):
        """Write a full element (typed instance or dict) at the given index."""
        if instance is not None:
            new_body = self._encoder.encode(instance)
        elif data is not None:
            new_body = self._encoder.encode_by_name(self._type_name, data)
        else:
            raise ValueError("Provide either 'instance' or 'data'.")
        self._write_body(index, new_body)

    @contextmanager
    def edit(self, index):
        """Context manager to modify an element in-place."""
        body = self._raw_values[index].Body if hasattr(self._raw_values[index], 'Body') else self._raw_values[index]
        instance = self._encoder.decode(self._type_name, body)
        yield instance
        new_body = self._encoder.encode(instance)
        self._write_body(index, new_body)

    def _write_body(self, index, new_body):
        """Write a single re-encoded element back to the server node."""
        new_values = []
        for i, raw in enumerate(self._raw_values):
            if i == index:
                new_values.append(ua.ExtensionObject(TypeId=raw.TypeId, Body=new_body))
            else:
                new_values.append(raw)
        self._node.set_value(new_values)
        self._refresh()
